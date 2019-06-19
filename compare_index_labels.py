import argparse
from openpyxl import load_workbook
import os
from ErrorLib import *


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument(
        '-b', '--baseIndexSheet', dest='base_index_sheet', required=True,
        help="Path to the old sample sheet with index padding"
    )
    parser.add_argument(
        '-c', '--compareIndexSheet', dest='compare_index_sheet', required=True,
        help="Path to new sample sheet with no index padding"
    )

    args = parser.parse_args()

    base_index_sheet = args.base_index_sheet
    compare_index_sheet = args.compare_index_sheet

    base_index_csv = make_csv(base_index_sheet)
    compare_index_csv = make_csv(compare_index_sheet)

    base_index_information = parse_index_sheet(base_index_csv)
    compare_index_information = parse_index_sheet(compare_index_csv)
    print(base_index_information)
    print(compare_index_information)

    for (index_name, index_sequence) in base_index_information.items():
        if index_sequence == compare_index_information[index_name]:
            print("Index " + str(index_name) + " with index sequence " + index_sequence +
                  " is the same between both files")
        else:
            print("ERROR! Index " + str(index_name) + " is not the same between two files!!")

    # Remove the csv files
    os.remove(base_index_csv)
    os.remove(compare_index_csv)


def make_csv(index_sheet_path):
    """
    Copy the extent of the xlsx file to a csv file with the same name
    :param index_sheet_path:
    :return:
    """
    file_name, ext = os.path.splitext(index_sheet_path)
    new_file_name = file_name + ".csv"
    index_csv = open(new_file_name, 'w')
    wb = load_workbook(index_sheet_path, read_only=True, data_only=True)
    worksheet = wb["Labels"]
    for row in worksheet.iter_rows():
        row_info = []
        for cell in row:
            row_info.append(str(cell.value))
        # Only concatenate the list with "," if is is not None
        index_csv.write(",".join(item for item in row_info if item))
        index_csv.write("\n")
    index_csv.close()
    return new_file_name


def parse_index_sheet(index_sheet_path):
    """
    Parse the csv file and gather the index name and sequence
    :param index_sheet_path:
    :return:
    """
    index_sheet = open(index_sheet_path, 'r')
    index_information = {}
    for line in index_sheet:
        if line.startswith("<LABEL ENTRIES>"):
            for line in index_sheet:
                # Exit the loop when the end of index section is reached
                if line.startswith("</LABEL ENTRIES>"):
                    index_sheet.close()
                    return index_information
                line = line.rstrip()
                line_item = line.split(",")
                label_name = line_item[2]
                index_sequence = line_item[3]
                if label_name not in index_information.keys():
                    index_information[label_name] = index_sequence
                elif label_name in index_information.keys():
                    index_sheet.close()
                    raise DuplicateIndexError("Duplicate index name found for " + label_name)


if __name__ == "__main__":
    main()
